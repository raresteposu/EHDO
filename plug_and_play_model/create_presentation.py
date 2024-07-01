import openpyxl
import json

building = "ac_sanierterzustand"

path_to_results_excel = "./results.xlsx"
wb = openpyxl.Workbook()
wb.create_sheet("Results")
wb.create_sheet("Data")
del wb['Sheet']

data_s = wb["Data"] # Data sheet
results_s = wb["Results"] # Results sheet

# Load JSON results
path_to_results_json = "./results/" + building + ".json"
with open(path_to_results_json, "r") as file:
    results_json = json.load(file)

# Initialize generation and cost dictionaries
generation_devices = {}
cost_devices = {}

# Process each device in the JSON
index_device = 1
for device in results_json["Devices"]:
    generation_devices[device] = {"Electricity": None, "Heating": None, "Cooling": None}
    
    if device in ["STC", "EB", "BBOI", "BOI"]:
        generation_devices[device]["Heating"] = results_json["Devices"][device]["generated"] if results_json["Devices"][device]["generated"] > 0 else None
    elif device == "PV":
        generation_devices[device]["Electricity"] = results_json["Devices"][device]["generated"]
    elif isinstance(results_json["Devices"][device]["generated"], dict):
        if "cool" in results_json["Devices"][device]["generated"]:
            generation_devices[device]["Cooling"] = results_json["Devices"][device]["generated"]["cool"] if results_json["Devices"][device]["generated"]["cool"] > 0 else None
        if "heat" in results_json["Devices"][device]["generated"]:
            generation_devices[device]["Heating"] = results_json["Devices"][device]["generated"]["heat"] if results_json["Devices"][device]["generated"]["heat"] > 0 else None
        if "power" in results_json["Devices"][device]["generated"]:
            generation_devices[device]["Electricity"] = results_json["Devices"][device]["generated"]["power"] if results_json["Devices"][device]["generated"]["power"] > 0 else None
    
    cost_devices[device] = results_json["Devices"][device]["cost"]

    # Ensure we are writing the correct data types
    data_s["A" + str(1 + index_device)] = device
    data_s["B" + str(1 + index_device)] = generation_devices[device]["Electricity"]
    data_s["C" + str(1 + index_device)] = generation_devices[device]["Heating"]
    data_s["D" + str(1 + index_device)] = generation_devices[device]["Cooling"]
    data_s["E" + str(1 + index_device)] = cost_devices[device]

    index_device += 1

# Populate costs
data_s["H2"] = float(results_json["Total Costs"]["Total device costs"]["Total investment costs"])
data_s["H3"] = float(results_json["Total Costs"]["Total device costs"]["Total O&M costs"])
data_s["H4"] = float(results_json["Total Costs"]["CO2 Tax"]) if float(results_json["Total Costs"]["CO2 Tax"]) > 0 else None
data_s["H5"] = (float(results_json["Total Costs"]["Supply costs"]["Electricity"]["Supply costs"]) +
                float(results_json["Total Costs"]["Supply costs"]["Electricity"]["Cap costs"]) -
                float(results_json["Total Costs"]["Supply costs"]["Electricity"]["Feed-in revenues"]))
data_s["H5"] = data_s["H5"].value if data_s["H5"].value > 0 else None
data_s["H6"] = (float(results_json["Total Costs"]["Supply costs"]["Gas"]["Supply costs"]) +
                float(results_json["Total Costs"]["Supply costs"]["Gas"]["Cap costs"]) -
                float(results_json["Total Costs"]["Supply costs"]["Gas"]["Feed-in revenues"]))
data_s["H6"] = data_s["H6"].value if data_s["H6"].value > 0 else None
data_s["H7"] = (float(results_json["Total Costs"]["Supply costs"]["Heat"]["Supply costs"]) +
                float(results_json["Total Costs"]["Supply costs"]["Heat"]["Cap costs"]))
data_s["H7"] = data_s["H7"].value if data_s["H7"].value > 0 else None
data_s["H8"] = float(results_json["Total Costs"]["Supply costs"]["Biomass"])
data_s["H8"] = data_s["H8"].value if data_s["H8"].value > 0 else None

# Populate CO2 emissions
data_s["K2"] = float(results_json["CO2 Emissions"]["Onsite"]) if float(results_json["CO2 Emissions"]["Onsite"]) > 0 else None
data_s["K3"] = float(results_json["CO2 Emissions"]["Generated due to electricity import"]) if float(results_json["CO2 Emissions"]["Generated due to electricity import"]) > 0 else None
data_s["K4"] = float(results_json["CO2 Emissions"]["Generated due to gas import"]) if float(results_json["CO2 Emissions"]["Generated due to gas import"]) > 0 else None
data_s["K5"] = float(results_json["CO2 Emissions"]["Generated due to heat import"]) if float(results_json["CO2 Emissions"]["Generated due to heat import"]) > 0 else None
data_s["K6"] = float(results_json["CO2 Emissions"]["Generated due to biom import"]) if float(results_json["CO2 Emissions"]["Generated due to biom import"]) > 0 else None
data_s["K8"] = float(results_json["CO2 Emissions"]["Won due to electricity export"]) if float(results_json["CO2 Emissions"]["Won due to electricity export"]) > 0 else None
data_s["K9"] = float(results_json["CO2 Emissions"]["Won due to gas export"]) if float(results_json["CO2 Emissions"]["Won due to gas export"]) > 0 else None

# Populate demands
results_s["O2"] = float(results_json["Demands"]["sum"]["power"])
results_s["O3"] = float(results_json["Demands"]["sum"]["heat"])
results_s["O4"] = float(results_json["Demands"]["sum"]["cool"]) if float(results_json["Demands"]["sum"]["cool"]) > 0 else None

results_s["O6"] = sum(generation_devices[device]["Electricity"] for device in generation_devices if generation_devices[device]["Electricity"] is not None)
results_s["O7"] = sum(generation_devices[device]["Heating"] for device in generation_devices if generation_devices[device]["Heating"] is not None)
results_s["O8"] = sum(generation_devices[device]["Cooling"] for device in generation_devices if generation_devices[device]["Cooling"] is not None)

results_s["O10"] = float(results_json["Grid Flows"]["Total electricity import"]) if float(results_json["Grid Flows"]["Total electricity import"]) > 0 else None
results_s["O11"] = float(results_json["Grid Flows"]["Total heat import"]) if float(results_json["Grid Flows"]["Total heat import"]) > 0 else None
results_s["O13"] = float(results_json["Grid Flows"]["Total electricity export"]) if float(results_json["Grid Flows"]["Total electricity export"]) > 0 else None

# Populate total annualized costs
results_s["O18"] = float(results_json["Total Costs"]["Total annualized costs"])

# Populate total imports and exports

results_s["O24"] = float(results_json["Grid Flows"]["Total electricity import"]) if float(results_json["Grid Flows"]["Total electricity import"]) > 0 else None
results_s["O25"] = float(results_json["Grid Flows"]["Total gas import"]) if float(results_json["Grid Flows"]["Total gas import"]) > 0 else None
results_s["O26"] = float(results_json["Grid Flows"]["Total biomass import"]) if float(results_json["Grid Flows"]["Total biomass import"]) > 0 else None

results_s["O28"] = float(results_json["Grid Flows"]["Total electricity export"]) if float(results_json["Grid Flows"]["Total electricity export"]) > 0 else None
results_s["O29"] = float(results_json["Grid Flows"]["Total gas export"]) if float(results_json["Grid Flows"]["Total gas export"]) > 0 else None


# Populate the data file text

data_s["A1"] = "Devices"
data_s["B1"] = "Power"
data_s["C1"] = "Heating"
data_s["D1"] = "Cooling"
data_s["E1"] = "Cost"

data_s["G1"] = "Cost Type"
data_s["G2"] = "Investment"
data_s["G3"] = "O&M"
data_s["G4"] = "CO2 Tax"
data_s["G5"] = "Electricity import"
data_s["G6"] = "Heating import"
data_s["G7"] = "Gas import"
data_s["G8"] = "Biomass import"

data_s["H1"] = "Cost"

data_s["J1"] = "CO2 Type"
data_s["J2"] = "Onsite"
data_s["J3"] = "Electricity import"
data_s["J4"] = "Gas import"
data_s["J5"] = "Biomass import"
data_s["J6"] = "Gas import"
data_s["J8"] = "Electricity export"
data_s["J9"] = "Gas export"
data_s["J11"] = "Tota generated"
data_s["J12"] = "Total credit won"

try:
    data_s["K11"] = data_s["K2"].value + data_s["K3"].value + data_s["K4"].value + data_s["K5"].value + data_s["K6"].value
except TypeError:
    data_s["K11"] = None
try:
    data_s["K12"] = data_s["K8"].value + data_s["K9"].value
except TypeError:
    data_s["K12"] = None




# Save the workbook
new_path = "./results/" + building + "_presentation.xlsx"
wb.save(new_path)

print("Workbook saved successfully.")
